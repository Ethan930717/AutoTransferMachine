import re
import cloudscraper
from bs4 import BeautifulSoup
import openpyxl
import time
from requests.cookies import cookiejar_from_dict
start_time = time.time()
#更换自己的cookie
cookie = ""
site = "shadowflow.org" #更换站点域名
def cookies_raw2jar(raw_cookies):
    cookie_dict = {}
    for cookie in raw_cookies.split(";"):
        key, value = cookie.split("=", 1)
        cookie_dict[key] = value
    return cookiejar_from_dict(cookie_dict)
scraper = cloudscraper.create_scraper()
wb = openpyxl.Workbook()  
ws = wb.active  
row = 2 
ws.title = f"{site}_torrents"
siteurl = f"https://{site}/"
for page in range(99999):
        torrent_url= f"{siteurl}torrents.php?page={page}"
        r = scraper.get(torrent_url, cookies=cookies_raw2jar(cookie),timeout=30)
        soup = BeautifulSoup(r.content, "html.parser")
        if r.status_code == 200:
            # 查找id为torrents的表格元素
            table = soup.find("table", class_="torrents")
            if table:
                trs = table.find_all("tr")[1:]
                ws["A1"] = "标题"
                ws["B1"] = "体积"
                ws["C1"] = "做种人数"
                ws["D1"] = "发布时间"
                ws["E1"] = "种子ID"
                ws["F1"] = "下载链接"
                for tr in trs: 
                    if "禁转" in tr.text or "分集" in tr.text :  # 这里可以排除不想爬的种子标签
                        continue  
                    try:
                        embedded = tr.find("td", class_="embedded")
                        a = embedded.find("a", href=lambda x: "details" in x)
                        b = a["href"]
                        title = a["title"]
                        details = f"{siteurl}{b}"
                        pattern = "id=(\d+)&hit"
                        torrent_id= re.search(pattern, details)
                        if torrent_id:
                            print(torrent_id.group(1)) 
                        else:
                            print("未识别到种子ID")
                        download = details.replace("details", "download")
                        download = download.replace("hit=1", "passkey=")
                        seeders = tr.find_all("td")[-4].text
                        size = tr.find_all("td")[-5].text
                        uploadtime = tr.find_all("td")[-6].text
                        ws["A" + str(row)] = title
                        ws["B" + str(row)] = size
                        ws["C" + str(row)] = seeders
                        ws["D" + str(row)] = uploadtime
                        ws["E" + str(row)] = torrent_id.group(1)
                        ws["F" + str(row)] = download
                        row += 1  # 行号加一

                    except IndexError:
                        print("啥也不是")
                    print(title,size,seeders,uploadtime,details)
            else:
                print("没东西了，停")
                break
        else:
            print("没东西了，停")
            continue
wb.save(f"{site}_torrents.xlsx")
total_rows = row - 1
total_pages = page + 1
end_time = time.time()
execution_time = end_time - start_time

print(f"爬取结束，本次共读取到{total_rows}个种子,耗时{execution_time}")