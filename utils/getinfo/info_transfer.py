import lxml
import cloudscraper
from bs4 import BeautifulSoup
from lxml import etree
import requests
from requests.cookies import cookiejar_from_dict
from AutoTransferMachine.utils.getinfo.makeyaml import mkyaml
import openpyxl
import urllib
from loguru import logger

def cookies_raw2jar(raw_cookies): # 定义一个函数，将原始的cookie字符串转换为cookiejar对象
    cookie_dict = {}
    for cookie in raw_cookies.split(";"):
        key, value = cookie.split("=", 1)
        cookie_dict[key] = value
    return cookiejar_from_dict(cookie_dict) # 调用requests模块中的函数
scraper = cloudscraper.create_scraper()


def getmediainfo(yamlinfo):
    wb = openpyxl.load_workbook(yamlinfo['basic']['torrent_list'])
    ws = wb.active
    data = ws.values
    url_list = [cell for row in data for cell in row if "detail" in cell]
    writemode = input(f"请选择模板转换方式\nY.在原有的pathinfo下自动续写\nN.覆盖原有的pathinfo，从path1开始生成（默认自动续写）")
    if writemode.lower() == "n":
        logger.info('当前为覆盖模式')
        f = open("au", "w", encoding="utf-8")
        lines = f.readlines()
        for index, line in enumerate(lines):
            if "path info" in line:
                path_index = index
                break
        new_lines = lines[:path_index + 1]
        for new_line in new_lines:
            f.write(new_line + "\n")
        f.close()
    else:
        logger.info('当前为续写模式')
    print(yamlinfo['basic']['torrent_list'])
    tmdb_api = yamlinfo['basic']['tmdb_api']
    counter = 0
    for url in url_list:
        result = urllib.parse.urlparse(url)
        siteurl = urllib.parse.urlunparse((result.scheme, result.netloc, '', '', '', ''))
        sitename = ws["G" + str(counter + 2)].value
        cookie = ws["H" + str(counter + 2)].value
        print(f"当前域名 {siteurl},匹配站点{sitename}")
        r = scraper.post(url, cookies=cookies_raw2jar(cookie), timeout=30)
        soup = BeautifulSoup(r.text, "html.parser")
        tree = lxml.etree.HTML(r.text)
        print(f"当前检索链接 {url}")

    #主标题
        xpath_name = "//*[@id='top']/text()"
        try:
            name = tree.xpath(xpath_name)[0]
            name = name.rstrip()
            getteam = name.split("-")
            # 取列表的最后一个元素，即最后一个"-"后面的内容
            team = getteam[-1]
        except IndexError:
            name = input(f"无法读取主标题名，请手动输入,种子地址{url}")
        print(f"成功记录主标题名 {name}")

    # 种子名称
        try:
            links = soup.find_all("a", class_="index")
            for link in links:
                href = link.get("href")
                if "download" in href:
                    torrent = link.text
                    filename=torrent.replace(".torrent","")
                    filename=filename.replace("[Shadow].","")
        except IndexError:
            torrent= input(f"无法读取种子名称，请手动输入,种子地址{url}")
        print(f"成功读取文件名称 {filename}")

        #副标题
        try:
            elements = soup.find_all("td", class_="rowfollow", valign="top", align="left")[0]
            texts_sdescr = [element.get_text() for element in elements]
            for small_descr in texts_sdescr:
               print(f"成功读取副标题名 {small_descr}")
        except IndexError:
            small_descr = input(f"无法读取副标题名，请手动输入,种子地址{url}")


        #产地&年份
        kdescr = soup.find(id="kdescr")
        lines = kdescr.text.split("\n")
        try:
            for line in lines:
                if "◎产　　地　" in line:
                    country = line.split("◎产　　地　")[1]
            print(f"读取产地成功 {country}")
        except IndexError:
            country= input(f"无法确认产地，请手动输入,文件标题{name}")

        try:
            for line in lines:
                if "◎年　　份　" in line:
                    date = line.split("◎年　　份　")[1]
            print(f"读取年份成功 {date}")
        except IndexError:
            date= input(f"无法确认年份，请手动输入,文件标题{name}")

        #标签
        try:
            elements = soup.find_all("td", class_="rowfollow", valign="top", align="left")[1]
            texts_tags = [element.get_text() for element in elements]
            tags = " ".join(texts_tags)
            print(f"读取标签成功 {tags}")
            if "mv" in tags.lower() or "体育" in tags or "音轨" in tags or "sport" in tags.lower() :
                print("当前尚未适配转载该类型资源，即将跳过本资源")
                continue
        except IndexError:
            tags = input(f"无法确认标签，请手动输入,资源链接{url}")

        #确认完结
        if "complete" in name.lower() and "完结" in tags:
            complete = 1
        elif "complete" in name.lower() and not "完结" in tags:
            choice = input("标题包含Complete，但资源未勾选完结标签，请手动确认该资源是否完结。输入'1'代表完结，输入'0'代表未完结")
            if choice == '1':
                complete = 1
            elif choice == '0':
                complete = 0
            else:
                print("无效的输入，请重新输入")
        elif not "complete" in name.lower() and "完结" in tags:
            choice = input("资源已勾选完结标签，但标题不包含Complete，请手动确认该资源是否完结。输入'1'代表完结，输入'0'代表未完结")
            if choice == '1':
                complete = 1
                name = input(f"请在主标题中手动加入Complete,添加位置在季数之后，如 S01 Complete\n当前主标题{name}")
            elif choice == '0':
                complete = 0
            else:
                print("无效的输入，请重新输入")
        else:
            complete = 0

        #基本信息
        try:
            elements = soup.find_all("td", class_="rowfollow", valign="top", align="left")[2]
            texts_info = [element.get_text() for element in elements]
            info = " ".join(texts_info)
            print("获取基本信息成功")
            infolist = info.split()
            size = " ".join(infolist[1:3])
            type = infolist[4]
            medium = infolist[6]
            codec = infolist[8]
            audio = infolist[10]
            standard = infolist[12]
            print(f"资源体积 {size}")
            print(f"类型 {type}")
            if "电影" in type:
                type = "movie"
            elif "剧集" in type:
                type = "tvseries"
            elif "综艺" in type:
                type = "tvshow"
            elif "纪录" in type:
                type = "doc"
            elif "动画" in type:
                type = "anime"
            elif "音轨" in type:
                type = "music"
                print("检测到类型为音轨，当前尚未适配转载该类型资源，即将跳过本资源")
                continue
            elif "MV" in type:
                type = "mv"
                print("检测到类型为MV，当前尚未适配转载该类型资源，即将跳过本资源")
                continue
            elif "体育" in type:
                type = "sport"
                print("检测到类型为体育，当前尚未适配转载该类型资源，即将跳过本资源")
                continue
            else:
                type = "other"
                print("未检测到当前资源类型，即将跳过本资源")
                continue

            print(f"媒介 {medium}")
            print(f"编码 {codec}")
            print(f"音频编码 {audio}")
            print(f"清晰度 {standard}")
            print(f"制作组 {team}")
        except IndexError:
            print("无法获取基本信息")

    # 豆瓣
        try:
            dblinks = soup.find_all("a", href=lambda x: x and "douban" in x)
            douban = [(link.get("href"), link.get_text()) for link in dblinks]
            for douban in dblinks:
                douban = douban.get("href")
                print(f"成功获取豆瓣链接 {douban}")
        except IndexError:
            douban = ""
            print("无法获取豆瓣链接")

        #IMDB
        imdb = ""
        try:
            imdblinks = soup.find_all("a", href=lambda x: x and "imdb" in x)
            if imdblinks:
                imdb = [(link.get("href"), link.get_text()) for link in imdblinks]
                for imdb in imdblinks:
                    imdb = imdb.get("href")
                    print(f"成功获取IMDB链接 {imdb}")
                    imdb_split=imdb.split('/')
                    imdb_id=imdb_split[4]
                    print(f"正在通过IMDBID '{imdb_id}' 获取TMDBID")
                    tmdb_url = "https://api.themoviedb.org/3/find/"+imdb_id+"?api_key="+tmdb_api+"&external_source=imdb_id&include_adult=true&language=zh-CN"
                    print(tmdb_url)
                    try:
                        tmdb_res = requests.get(tmdb_url)
                        if tmdb_res.status_code == 200:
                            data = tmdb_res.json()
                            if data["movie_results"]:
                                tmdb_id = data["movie_results"][0]["id"]
                                tmdb_title = data["movie_results"][0]["title"]
                                print(f"TMDBID获取成功 {tmdb_id} ")
                                print(f"中文标题 {tmdb_title} ")
                            elif data["tv_results"]:
                                tmdb_id = data["tv_results"][0]["id"]
                                tmdb_title = data["tv_results"][0]["name"]
                                print(f"TMDBID获取成功 {tmdb_id} ")
                                print(f"中文标题 {tmdb_title} ")
                            elif data["tv_episode_results"]:
                                tmdb_id = data["tv_episode_results"][0]["id"]
                                tmdb_title = data["tv_episode_results"][0]["name"]
                                print(f"TMDBID获取成功 {tmdb_id} ")
                                print(f"中文标题 {tmdb_title} ")
                            else:
                                print("未成功获取TMDBID，尝试使用影片名称搜索")
                                if type == "movie":
                                    print("")
                                else:
                                    print("")
                        else:
                            print(f"TMDB请求失败 {tmdb_res.status_code}.")
                    except IndexError:
                        tmdb_id = ""
                        tmdb_title = ""
                        print("无法连接到tmdb")
                else:
                    imdb = ""
                    print("该资源暂无imdb链接")
        except Exception as e:
            print("无法获取IMDB链接")
        logger.info(f"第{counter}个资源读取完成")
        return mkyaml(yamlinfo,counter, filename, name, small_descr, tags, team, type, audio, codec, medium, douban, imdb, country, date, standard, tmdb_id,  torrent)





