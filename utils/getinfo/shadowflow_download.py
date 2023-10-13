import re
import cloudscraper
from bs4 import BeautifulSoup
import openpyxl
from requests.cookies import cookiejar_from_dict
from loguru import logger
import time
from loguru import logger
import sys
import shutil
import yaml
import urllib
import requests
import os
import logging
from qbittorrentapi import Client
from utils.getinfo.torrent_download import *

def cookies_raw2jar(raw_cookies):
    cookie_dict = {}
    for cookie in raw_cookies.split(";"):
        key, value = cookie.split("=", 1)
        cookie_dict[key] = value
    return cookiejar_from_dict(cookie_dict)

def shadowflow_down(raw_cookies,sitename,siteurl,sitecookie,sitepasskey,yamlinfo,start_time):
    scraper = cloudscraper.create_scraper()
    wb = openpyxl.Workbook()
    ws = wb.active
    row = 2
    ws.title = f"{sitename}_torrents"
    try:
        pagenum = int(input('请输入本次需要爬取几页种子: '))
        if pagenum > 0:  # 如果输入是正整数
            print(f"您选择了爬取{pagenum}页种子")
    except ValueError:
        print('输入错误，请输入一个整数。')
    while True:

        outtag = input(
            f"请选择需要排除的资源关键字，可多选，无格式要求（默认排除禁转、限转资源，注意区分大小写) \n例：排除有国语粤语标签的动漫和综艺资源，则输入CD12 \n做种人数(seed)，体积(size)筛选请使用前后比较符，如0<seed<5则为排除做种人数小于5人的种子，10<size<100则表示排除体积在10GB-100GB之间的种子，可与关键字一同筛选）\n A.电影 B.剧集 C.综艺 D.动漫 E.纪录片 F.MV\n gy.国语 yy.粤语 zz.中字 diy.DIY wj.完结 fj.分集 db.杜比视界 hdr.HDR\n请输入排除项:")
        tags = []
        tags.append("禁转")
        tags.append("限转")
        if "A" in outtag:
            tags.append("电影")
        if "B" in outtag:
            tags.append("剧集")
        if "C" in outtag:
            tags.append("综艺")
        if "D" in outtag:
            tags.append("动漫")
        if "E" in outtag:
            tags.append("纪录片")
        if "F" in outtag:
            tags.append("MV")
        if "gy" in outtag:
            tags.append("国语")
        if "yy" in outtag:
            tags.append("粤语")
        if "zz" in outtag:
            tags.append("中字")
        if "diy" in outtag:
            tags.append("DIY")
        if "wj" in outtag:
            tags.append("完结")
        if "fj" in outtag:
            tags.append("分集")
        if "db" in outtag:
            tags.append("杜比视界")
        if "hdr" in outtag:
            tags.append("HDR")
        tags_str = " ".join(tags)
        # 解析输入，查找是否包含seed筛选条件
        seed_filter = None
        seedprint = ""
        if 'seed' in outtag:
            match = re.search(r'(\d+)\s*<\s*seed\s*<\s*(\d+)', outtag)
            if match:
                seed_min = int(match.group(1))
                seed_max = int(match.group(2))
                seed_filter = lambda x: seed_min <= int(x) < seed_max
                seedprint = f"做种人数在{seed_min}到{seed_max}之间,"

        # 解析输入，查找是否包含size筛选条件
        size_filter = None
        sizeprint = ""
        if 'size' in outtag:
            match = re.search(r'(\d+)\s*<\s*size\s*<\s*(\d+)', outtag)
            if match:
                size_min = int(match.group(1))
                size_max = int(match.group(2))
                size_filter = lambda x: size_min <= int(x) < size_max
                sizeprint = f"体积在{size_min}GB到{size_max}GB之间"
        download_ensure = input(f"选择完毕，本次将为您排除{tags_str},{seedprint}{sizeprint}的资源\n确认：Y \n重选： N \n")
        if download_ensure.lower() == "y":
            break
        elif download_ensure.lower() == "n":
            continue
        else:
            print("输入有误，请输入Y或者N")
    for page in range(pagenum):
        torrent_url = f"{siteurl}torrents.php?page={page}"
        r = scraper.get(torrent_url, cookies=cookies_raw2jar(sitecookie), timeout=30)
        soup = BeautifulSoup(r.content, "html.parser")
        if r.status_code == 200:
            # 查找id为torrents的表格元素
            table = soup.find("table", class_="torrents")
            if table:
                trs = table.find_all("tr")[1:]
                ws["A1"] = "标题"
                ws["B1"] = "体积"
                ws["C1"] = "做种人数"
                ws["D1"] = "发布时间"
                ws["E1"] = "种子ID"
                ws["F1"] = "下载链接"
                ws["G1"] = "站点"
                ws["H1"] = "cookie"
                ws["I1"] = "passkey"
                for tr in trs:
                    if any(x in tr.text for x in tags):
                        print(f"不符合筛选条件，跳过")
                        continue
                    else:
                        td_list = tr.find_all("td")
                        if len(td_list) >= 4:
                            seeders = td_list[-4].text
                        else:
                            continue
                        size = tr.find_all("td")[-5].text
                        size_pattern = re.search(r'(\d+\.\d+|\d+)', size)
                        size = float(size_pattern.group(1)) / 1024 if 'MB' in size else float(size_pattern.group(1))
                        # 将MB单位的体积转换为GB单位
                        if (seed_filter is None or seed_filter(seeders)) and (size_filter is None or size_filter(size)):
                            try:
                                embedded = tr.find("td", class_="embedded")
                                a = embedded.find("a", href=lambda x: "details" in x)
                                b = a["href"]
                                title = a["title"]
                                details = f"{siteurl}{b}"
                                pattern = "id=(\d+)&hit"
                                torrent_id = re.search(pattern, details)
                                if torrent_id:
                                    pass
                                else:
                                    print("未识别到种子ID")
                                download = details.replace("details", "download")
                                download = download.replace("hit=1", f"passkey={sitepasskey}")
                                seeders = tr.find_all("td")[-4].text
                                size = tr.find_all("td")[-5].text
                                uploadtime = tr.find_all("td")[-6].text
                                ws["A" + str(row)] = title
                                ws["B" + str(row)] = size
                                ws["C" + str(row)] = seeders
                                ws["D" + str(row)] = uploadtime
                                ws["E" + str(row)] = details
                                # ws["E" + str(row)] = torrent_id.group(1)
                                ws["F" + str(row)] = download
                                ws["G" + str(row)] = sitename
                                ws["H" + str(row)] = sitecookie
                                ws["I" + str(row)] = sitepasskey
                                row += 1  # 行号加一
                                logger.info(f'{title}获取成功')
                            except IndexError:
                                continue
                        else:
                            print(f"不符合筛选条件，跳过")
            else:
                print("没东西了，停")
                break
        else:
            print("没东西了，停")
            continue
    wb.save(f"{sitename}_torrents.xlsx")
    total_rows = row - 1
    total_pages = page + 1
    end_time = time.time()
    execution_time = end_time - start_time
    logger.info(f"爬取结束，本次共读取到{total_rows}个种子,耗时{execution_time}，请选择接下来的任务\n 1.批量打印种子链接 2.批量打印下载链接 3.跳过")
    shutil.move(yamlinfo['basic']['screenshot_path'] + '/' + sitename + '_torrents.xlsx',yamlinfo['basic']['record_path'] + '/' + sitename + '_torrents.xlsx')
    choice = input("请输入您的选择：")
    while True:
        if choice == "1":
            print("以下是所有的种子链接：")
            for i in range(2, row):
                details = ws["E" + str(i)].value
                print(details)
            break
        elif choice == "2":
            print("以下是所有的下载链接：")
            for i in range(2, row):
                download = ws["F" + str(i)].value
                print(download)
            break
        elif choice == "3":
            print("好的")
            break
        else:
            logger.info("选择错误，请重新选择")
            continue
    #获取path序列
    while True:
        forsure = input(f"本次数据已保存在{yamlinfo['basic']['record_path']}/{sitename}_torrents.xlsx\n是否需要将Yaml模板中的torrent_file路径替换成本次生成的数据文件路径\nY.是，替换路径\nN.否，不需要替换\n默认不替换")
        if forsure.upper() == 'Y':
            au = f"{yamlinfo['basic']['workpath']}au.yaml"
            logger.info(f"检测到模板路径为{au}")
            with open(au, "r") as f:
                yamlinfo = yaml.load(f, Loader=yaml.FullLoader)
                yamlinfo["basic"]["torrent_list"] = f"{yamlinfo['basic']['record_path']}/{sitename}_torrents.xlsx"
                with open(au, "w") as f:
                    yaml.dump(yamlinfo, f)
            logger.info(f"修改完成，当前yaml模板的torrent_list为{yamlinfo['basic']['torrent_list']}")
            while True:
                dlsure = input(
                    f'是否需要将本次抓取到的资源种子下载到本地（下载路径为torrent_path，请确认配置文件中已正确配置该项\nY.是，下载\nN，否，不下载')
                if dlsure.upper() == "Y":
                    logger.info("开始下载")
                    return download_torrent(ws, yamlinfo)
                elif dlsure.upper() == "N":
                    logger.info("未选择下载种子，即将结束本次任务")
                    sys.exit()
                else:
                    logger.info("选择错误，请重新选择")
                    continue
        elif forsure.upper() == 'N':
            logger.info('已选择不替换路径')
            while True:
                dlsure = input(f'是否需要将本次抓取到的资源种子下载到本地（选是则同步将种子上传到QB，默认添加后为暂停状态）\nY.是，下载\nN，否，不下载')
                if dlsure.upper() == "Y":
                    logger.info("开始下载")
                    return download_torrent(ws, yamlinfo)
                elif dlsure.upper() == "N":
                    logger.info("未选择下载种子，即将结束本次任务")
                    sys.exit()
                else:
                    logger.info("选择错误，请重新选择")
                    continue
        else:
            logger.info("选择错误，请重新选择")
            continue